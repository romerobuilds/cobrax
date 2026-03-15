# app/routes/client.py
from __future__ import annotations

import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import List, Dict, Any, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.core.deps import get_current_user, get_company_for_current_user
from app.database_.database import get_db
from app.models.client import Client
from app.models.company import Company
from app.models.user import User
from app.schemas.client import ClientCreate, ClientPublic, ClientUpdate

from openpyxl import load_workbook


router = APIRouter(
    prefix="/empresas/{company_id}/clientes",
    tags=["Clientes"],
    dependencies=[Depends(get_company_for_current_user)],
)


def _get_client_or_404(db: Session, company_id: UUID, client_id: UUID) -> Client:
    client = (
        db.query(Client)
        .filter(
            Client.id == client_id,
            Client.company_id == company_id,
        )
        .first()
    )
    if not client:
        raise HTTPException(status_code=404, detail="Cliente não encontrado")
    return client


# -------------------------
# Upload helpers
# -------------------------

def _normalize_header(s: str) -> str:
    return (s or "").strip()


def _parse_csv(raw: bytes, limit: int) -> List[Dict[str, str]]:
    text = raw.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    rows: List[Dict[str, str]] = []
    for i, row in enumerate(reader):
        if i >= limit:
            break
        if not row:
            continue
        clean: Dict[str, str] = {}
        for k, v in row.items():
            hk = _normalize_header(k)
            if not hk:
                continue
            clean[hk] = "" if v is None else str(v).strip()
        if clean:
            rows.append(clean)
    return rows


def _parse_xlsx(raw: bytes, limit: int) -> List[Dict[str, str]]:
    wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return []

    headers_norm = [_normalize_header(str(h) if h is not None else "") for h in headers]
    rows: List[Dict[str, str]] = []

    for i, values in enumerate(rows_iter):
        if i >= limit:
            break
        row: Dict[str, str] = {}
        for idx, h in enumerate(headers_norm):
            if not h:
                continue
            v = values[idx] if idx < len(values) else None
            row[h] = "" if v is None else str(v).strip()
        if row:
            rows.append(row)

    return rows


def _parse_upload_file(file: UploadFile, raw: bytes, limit: int) -> List[Dict[str, str]]:
    filename = (file.filename or "").lower()
    ctype = (file.content_type or "").lower()

    if filename.endswith(".csv") or "csv" in ctype:
        return _parse_csv(raw, limit)

    if filename.endswith(".xlsx") or "spreadsheet" in ctype or "excel" in ctype:
        return _parse_xlsx(raw, limit)

    raise HTTPException(status_code=400, detail="Formato inválido. Envie CSV ou XLSX.")


def _find_value_ci(row: Dict[str, Any], col: str) -> Optional[str]:
    col = (col or "").strip()
    if not col:
        return None

    if col in row:
        return row.get(col)

    target = col.lower()
    for k in row.keys():
        if (k or "").strip().lower() == target:
            return row.get(k)
    return None


def _parse_bool(val: Optional[str]) -> Optional[bool]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s == "":
        return None
    if s in {"1", "true", "t", "yes", "y", "sim", "s"}:
        return True
    if s in {"0", "false", "f", "no", "n", "nao", "não"}:
        return False
    return None


def _parse_money(val: Optional[str]) -> Optional[Decimal]:
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None

    s = s.replace("R$", "").replace("r$", "").strip()
    s = s.replace(" ", "")

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")

    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _sanitize_cpf_cnpj(val: Optional[str]) -> Optional[str]:
    if val is None:
        return None
    s = re.sub(r"\D+", "", str(val).strip())
    if not s:
        return None
    if len(s) not in (11, 14):
        return None
    return s


# -------------------------
# CRUD
# -------------------------

@router.post("/", response_model=ClientPublic, status_code=status.HTTP_201_CREATED)
def criar_cliente(
    company_id: UUID,
    payload: ClientCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    existe = (
        db.query(Client)
        .filter(Client.company_id == company_id, Client.email == str(payload.email))
        .first()
    )
    if existe:
        raise HTTPException(status_code=400, detail="Já existe cliente com esse e-mail nesta empresa")

    client = Client(
        nome=payload.nome,
        email=str(payload.email),
        telefone=payload.telefone,
        cpf_cnpj=_sanitize_cpf_cnpj(getattr(payload, "cpf_cnpj", None)),
        owner_id=user.id,
        company_id=company_id,
        is_mensalista=bool(payload.is_mensalista or False),
        saldo_aberto=payload.saldo_aberto if payload.saldo_aberto is not None else Decimal("0.00"),
    )
    db.add(client)
    db.commit()
    db.refresh(client)
    return client


@router.get("/", response_model=List[ClientPublic])
def listar_clientes(
    company_id: UUID,
    db: Session = Depends(get_db),
):
    return (
        db.query(Client)
        .filter(Client.company_id == company_id)
        .order_by(Client.created_at.desc())
        .all()
    )


@router.get("/{client_id}", response_model=ClientPublic)
def obter_cliente(
    company_id: UUID,
    client_id: UUID,
    db: Session = Depends(get_db),
):
    return _get_client_or_404(db, company_id, client_id)


@router.put("/{client_id}", response_model=ClientPublic)
def atualizar_cliente(
    company_id: UUID,
    client_id: UUID,
    payload: ClientUpdate,
    db: Session = Depends(get_db),
):
    client = _get_client_or_404(db, company_id, client_id)

    if payload.email and str(payload.email) != client.email:
        existe = (
            db.query(Client)
            .filter(
                Client.company_id == company_id,
                Client.email == str(payload.email),
                Client.id != client.id,
            )
            .first()
        )
        if existe:
            raise HTTPException(status_code=400, detail="Já existe cliente com esse e-mail nesta empresa")

    if payload.nome is not None:
        client.nome = payload.nome
    if payload.email is not None:
        client.email = str(payload.email)
    if payload.telefone is not None:
        client.telefone = payload.telefone

    if hasattr(payload, "cpf_cnpj") and payload.cpf_cnpj is not None:
        client.cpf_cnpj = _sanitize_cpf_cnpj(payload.cpf_cnpj)
    elif hasattr(payload, "cpf_cnpj") and payload.cpf_cnpj is None:
        client.cpf_cnpj = None

    if payload.is_mensalista is not None:
        client.is_mensalista = bool(payload.is_mensalista)
    if payload.saldo_aberto is not None:
        client.saldo_aberto = payload.saldo_aberto

    db.commit()
    db.refresh(client)
    return client


@router.delete("/{client_id}", status_code=status.HTTP_204_NO_CONTENT)
def deletar_cliente(
    company_id: UUID,
    client_id: UUID,
    db: Session = Depends(get_db),
):
    client = _get_client_or_404(db, company_id, client_id)
    db.delete(client)
    db.commit()
    return None


# =========================
# UPLOAD (CSV/XLSX)
# =========================

@router.post("/upload", operation_id="clients_upload")
async def upload_clients_file(
    company_id: UUID,
    file: UploadFile = File(...),
    email_column: str = Query(default="email", description="Nome da coluna do e-mail"),
    nome_column: str = Query(default="nome", description="Nome da coluna do nome"),
    telefone_column: str = Query(default="telefone", description="Nome da coluna do telefone"),
    cpf_cnpj_column: str = Query(default="cpf_cnpj", description="Coluna cpf_cnpj (somente números)"),
    mensalista_column: str = Query(default="mensalista", description="Coluna mensalista (sim/nao, true/false, 1/0)"),
    saldo_column: str = Query(default="saldo_aberto", description="Coluna saldo (ex: 123,45)"),
    limit: int = Query(default=5000, ge=1, le=50000, description="Máximo de linhas lidas"),
    update_existing: bool = Query(default=False, description="Se true, atualiza cliente existente pelo e-mail"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Arquivo vazio")

    rows = _parse_upload_file(file, raw, limit)
    if not rows:
        raise HTTPException(status_code=400, detail="Nenhuma linha válida encontrada")

    email_col = (email_column or "email").strip()
    nome_col = (nome_column or "nome").strip()
    tel_col = (telefone_column or "telefone").strip()
    cpf_col = (cpf_cnpj_column or "cpf_cnpj").strip()
    mensalista_col = (mensalista_column or "mensalista").strip()
    saldo_col = (saldo_column or "saldo_aberto").strip()

    existing = {
        (e or "").lower(): cid
        for (e, cid) in db.query(Client.email, Client.id)
        .filter(Client.company_id == company_id)
        .all()
        if e
    }

    created = 0
    updated = 0
    skipped_no_email = 0
    skipped_duplicate = 0
    errors: List[str] = []

    for idx, row in enumerate(rows, start=1):
        email_val = _find_value_ci(row, email_col)
        email = (email_val or "").strip()
        if not email:
            skipped_no_email += 1
            continue

        email_key = email.lower()

        nome_val = _find_value_ci(row, nome_col)
        telefone_val = _find_value_ci(row, tel_col)
        cpf_val = _find_value_ci(row, cpf_col)
        mensalista_val = _find_value_ci(row, mensalista_col)
        saldo_val = _find_value_ci(row, saldo_col)

        nome = (nome_val or "").strip()
        telefone = (telefone_val or "").strip()

        cpf_clean = _sanitize_cpf_cnpj(cpf_val)
        mensalista_bool = _parse_bool(mensalista_val)
        saldo_dec = _parse_money(saldo_val)

        if email_key in existing:
            if not update_existing:
                skipped_duplicate += 1
                continue

            client_id = existing[email_key]
            c = (
                db.query(Client)
                .filter(
                    Client.id == client_id,
                    Client.company_id == company_id,
                )
                .first()
            )
            if not c:
                skipped_duplicate += 1
                continue

            if nome:
                c.nome = nome
            if telefone:
                c.telefone = telefone
            if cpf_clean:
                c.cpf_cnpj = cpf_clean
            if mensalista_bool is not None:
                c.is_mensalista = mensalista_bool
            if saldo_dec is not None:
                c.saldo_aberto = saldo_dec

            updated += 1
            continue

        if not nome:
            nome = email.split("@")[0]

        c = Client(
            nome=nome,
            email=email,
            telefone=telefone or None,
            cpf_cnpj=cpf_clean,
            owner_id=user.id,
            company_id=company_id,
            is_mensalista=mensalista_bool if mensalista_bool is not None else False,
            saldo_aberto=saldo_dec if saldo_dec is not None else Decimal("0.00"),
        )

        db.add(c)
        try:
            db.flush()
            existing[email_key] = c.id
            created += 1
        except IntegrityError:
            db.rollback()
            skipped_duplicate += 1
        except Exception as e:
            db.rollback()
            errors.append(f"linha {idx}: {str(e)}")

    db.commit()

    return {
        "ok": True,
        "created": int(created),
        "updated": int(updated),
        "skipped_no_email": int(skipped_no_email),
        "skipped_duplicate": int(skipped_duplicate),
        "errors": errors[:20],
        "note": "CSV/XLSX: email obrigatório. Campos opcionais: nome, telefone, cpf_cnpj, mensalista, saldo_aberto. update_existing=true atualiza por email.",
    }